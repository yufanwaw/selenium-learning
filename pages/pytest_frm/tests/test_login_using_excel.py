from pathlib import Path

import pytest
from openpyxl import load_workbook

from pages.login_page import LoginPage

LOGIN_URL = "https://the-internet.herokuapp.com/login"
PROJECT_ROOT = Path(__file__).resolve().parents[3]
EXCEL_FILE = PROJECT_ROOT / "data_source" / "login_data.xlsx"

VALID_USERNAME = "tomsmith"
VALID_PASSWORD = "SuperSecretPassword!"


def get_expected_message(username, password):
    """Decide the expected message for each username/password from Excel."""
    if username == VALID_USERNAME and password == VALID_PASSWORD:
        return "You logged into a secure area!"

    if username != VALID_USERNAME:
        return "Your username is invalid!"

    return "Your password is invalid!"


def load_login_data():
    """Read Excel rows and return data for pytest parametrize."""
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active
    test_data = []

    # Excel format:
    # username | password
    for username, password in sheet.iter_rows(min_row=2, values_only=True):
        if not username or not password:
            continue

        message = get_expected_message(username, password)
        test_data.append((username, password, message))

    return test_data


@pytest.mark.parametrize(
    "username,password,message",
    load_login_data(),
)
def test_login(driver, username, password, message):
    driver.get(LOGIN_URL)

    login_page = LoginPage(driver)
    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    assert message in login_page.get_flash_message()


# Run without HTML report:
# ./.venv/bin/python -m pytest pages/pytest_frm/tests/test_login_using_excel.py -v

# Run with HTML report:
# ./.venv/bin/python -m pytest pages/pytest_frm/tests/test_login_using_excel.py -v --html=report.html
